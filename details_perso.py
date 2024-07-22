import sys, subprocess
from PyQt5.QtWidgets import QApplication, QWidget, QDialog, QLabel, QPushButton, QLineEdit, QDateTimeEdit, QComboBox, QTableWidget, QTableWidgetItem
import pandas as pd
from PyQt5.QtCore import Qt, QPoint
from PyQt5.QtGui import QPixmap, QIcon
from openpyxl import Workbook
from fonction import insertion, ajout_fiche
from fpdf import FPDF
from base import *
db, connection = base()


class Tableau_fiche_ui(QWidget):
    def __init__(self, gestion_personnel):
        super().__init__()
        self.setWindowTitle("Gestion de Fiche")
        # self.setStyleSheet("background-color:#012b3c")
        self.resize(1500,1500)
        self.deconnecter = gestion_personnel


        self.imageFont =QLabel(self)
        self.setGeometry(0,0,1500,1500)
        imageFont = QPixmap("./images/image2.jpg")
        size = imageFont.scaled(1500,1500)
        self.imageFont.setPixmap(size)



        self.image =QLabel(self)
        self.image.setGeometry(30,30,70,70)
        self.image.setStyleSheet("border:2px solid white;border-radius:10px")
        image = QPixmap("./images/logo.png")
        size = image.scaled(70,70)
        self.image.setPixmap(size)

        self.titre_contenir = QLabel(self, text="DETAILS FICHES")
        self.titre_contenir.setGeometry(500,50,500,50)
        self.titre_contenir.setStyleSheet("font-size:40px;font-style:normal;font-weight:bold;color:#3c1101")

        self.annonce = QLabel(self)
        self.annonce.setGeometry(300,600,400,30)
        self.annonce.setText("Veuillez séléctionner une ligne dans le tableau pour faire l'action :")
        self.annonce.setStyleSheet("fonte-size:40px;color:#3c1101;font-weight:bold;text-decoration: underline")


        self.buton_imprimer = QPushButton(self)
        self.buton_imprimer.setGeometry(700,600,100,30)
        self.buton_imprimer.setText("IMPRIMER")
        self.buton_imprimer.setCursor(Qt.PointingHandCursor)
        self.buton_imprimer.setStyleSheet("background-color:blue;font-size:16px;color:white;border-radius:10px;font-weight:bold")
        self.buton_imprimer.enterEvent = self.ouvre
        self.buton_imprimer.leaveEvent = self.ferme
        self.buton_imprimer.setEnabled(False)
        
        self.buton_excel = QPushButton(self)
        self.buton_excel.setGeometry(950,600,100,30)
        self.buton_excel.setText("EXCEL")
        self.buton_excel.setCursor(Qt.PointingHandCursor)
        self.buton_excel.setStyleSheet("background-color:green;font-size:16px;color:white;border-radius:10px;font-weight:bold")
        self.buton_excel.enterEvent = self.ouvr
        self.buton_excel.leaveEvent = self.ferm
        self.buton_excel.setEnabled(False)
        
        
        
        self.buton_effacer = QPushButton(self)
        self.buton_effacer.setGeometry(830,600,100,30)
        self.buton_effacer.setText("SUPPRIMER")
        self.buton_effacer.setCursor(Qt.PointingHandCursor)
        self.buton_effacer.setStyleSheet("background-color:red;color:white;border-radius:10px;font-weight:bold")
        self.buton_effacer.enterEvent = self.ouvrir
        self.buton_effacer.leaveEvent = self.fermer
        self.buton_effacer.clicked.connect(self.Effacer)
        self.buton_effacer.setEnabled(False)

        # self.buton_effacer.setVisible(False)

        self.text = QLabel(self, text="selectionner dans le tableau pour effacer")
        self.text.setStyleSheet("""QLabel{color:black;font-size:12px;background-color:orange;padding:5px}""")
        self.text.setVisible(False)
        
        self.text1 = QLabel(self, text="selectionner un ligne dans le tableau pour imprimer")
        self.text1.setStyleSheet("""QLabel{color:black;font-size:12px;background-color:orange;padding:5px}""")
        self.text1.setVisible(False)

        self.text2 = QLabel(self, text="selectionner un ligne dans le tableau pour envoyer")
        self.text2.setStyleSheet("""QLabel{color:black;font-size:12px;background-color:orange;padding:5px}""")
        self.text2.setVisible(False)

        self.recherche_input = QLineEdit(self)
        self.recherche_input.setGeometry(500,150,200,30)
        self.recherche_input.setStyleSheet("font-size:20px;border-radius:10px;color:black;background-color:white")
        self.btn_recherche = QPushButton(self, text="recherche")
        self.btn_recherche.setGeometry(730,150,100,30)
        self.btn_recherche.setStyleSheet("font-size:20px;border-radius:10px;color:white;background-color:orange")
        self.btn_recherche.clicked.connect(self.Rechercher)


        self.tableau = QTableWidget(self)
        self.tableau.setGeometry(400,200,700,400)
        self.tableau.verticalHeader().setVisible(False)

        self.tableau.setStyleSheet("""
            QTableWidget {
                background-image: url(./images/image2.jpg);
                color: #012b3c;
                border-radius: 10px
            }
            QTableWidget::item {
                background-color: #8f028f77;
                border-bottom: 2px solid white;
                border-radius: 3px
            }  
            QTableWidget::item:selected {
                background-color: orange
            }
            
            QHeaderView::section {
                background-color: #3c1101;
                color: white;
                font-weight: bold;
                border: 1px solid white;
                border-radius: 5px   
            }
            QScrollBar:horizontal {
                background-color: #8f028f77;
                border: none;                
                height: 15px
            }
            QScrollBar::handle:horizontal {
                background-color: #012b3c;
                border-radius: 7px
            }
        """)

        self.tableau.setColumnCount(7)
        self.tableau.setRowCount(20)
        self.tableau.setHorizontalHeaderLabels([
           "id", "numero","date","nom", "poste", "conge", "evaluation"
        ])
        self.tableau.setColumnWidth(0,50)
        self.tableau.setColumnWidth(1,100)
        self.tableau.setColumnWidth(2,100)
        self.tableau.setColumnWidth(3,100)
        self.tableau.setColumnWidth(4,100)
        self.tableau.setColumnWidth(5,100)
        self.tableau.setColumnWidth(6,150)
        self.tableau.itemClicked.connect(self.Selectionner)
        
        listes=ajout_fiche()
        self.tableau.setRowCount(len(listes))
        for row, rowData in enumerate(listes):
            for col, value in enumerate(rowData):
                item = QTableWidgetItem(str(value))
                self.tableau.setItem(row,col,item)










        self.bouton_deconn = QPushButton(self)
        self.bouton_deconn.setGeometry(1200, 20, 50, 50)
        self.bouton_deconn.setIcon(QIcon('./images/close_icon.png'))
        self.bouton_deconn.setToolTip('Fermer l\'application')
        self.bouton_deconn.setCursor(Qt.PointingHandCursor)
        self.bouton_deconn.clicked.connect(self.retour_fenetre1)


        self.buton_imprimer.clicked.connect(self.generer_pdf)
        self.buton_excel.clicked.connect(self.print) 

    def Rechercher(self):
        mot= self.recherche_input.text()
        listes=ajout_fiche()
        self.tableau.setRowCount(len(listes))
        j=1
        for i in range(len(listes)):
            if mot in listes[i]:
                self.tableau.setRowCount(j)
                for col, value in enumerate(listes[i]):
                    item = QTableWidgetItem(str(value))
                    self.tableau.setItem(j-1,col,item)
                j += 1

    def ouvrir(self, event):
        self.text.setVisible(True)
        self.text.move(self.buton_effacer.pos() + QPoint(10, 20))
    def fermer(self, event):
        self.text.setVisible(False)

    def ouvre(self, event):
        self.text1.setVisible(True)
        self.text1.move(self.buton_imprimer.pos() + QPoint(10, 20))
    def ferme(self, event):
        self.text1.setVisible(False)

    def ouvr(self, event):
        self.text2.setVisible(True)
        self.text2.move(self.buton_excel.pos() + QPoint(10, 20))
    def ferm(self, event):
        self.text2.setVisible(False)



    def generer_pdf(self):
        # Récupérer l'indice de la ligne sélectionnée
        selected_row = self.tableau.currentRow()

        if selected_row == -1:
            QMessageBox.warning(self, "Aucune sélection", "Veuillez sélectionner une ligne.")
            return    

        # Créer un objet PDF
        pdf = FPDF()
        class PDF(FPDF):
            def header(self):
                self.set_font('Arial', 'B', 14)
                self.cell(0, 10, 'Fiche de renseignement', 0, 1, 'C')
                self.ln(10)

        pdf = PDF()
        pdf.add_page()
        pdf.set_font("Arial", size=12)

        labels = ["ID:", "Numéro:", "Date:", "Nom:", "Poste:", "Congé:", "Évaluation:"]


        donnees_tableau = []
        for row in range(self.tableau.rowCount()):
            ligne = []
            for col in range(self.tableau.columnCount()):
                item = self.tableau.item(selected_row, col)
                if item is not None:
                    ligne.append(labels[col] + " " + item.text())
            # donnees_tableau.append(ligne)

        contact_x = 130

        
        for value in ligne:
            pdf.cell(0, 10, value, ln=True)

        position_y = 20

        pdf.set_xy(contact_x, position_y)
        contact_info = """
        Contact :
        Adresse : [Route_Analamahitsy]
        Téléphone : [+025 025 055 055]
        E-mail : info@meublesplus.com
        Site web : www.meublesplus.com
        """
        pdf.set_font("Arial", size=10)
        pdf.multi_cell(0, 10, contact_info, align="R")

        pdf.cell(200, 10, ln=True)
        pdf.cell(200, 10, ln=True)
        pdf.cell(200, 10, ln=True)
        pdf.cell(200, 10, ln=True)

        pdf.set_font("Arial", size=12)
        mission_text = """
        Signature personnel :
        """
        pdf.multi_cell(0, 10, mission_text)


        # Sauvegarder le PDF
        pdf_output = "tableau_fiche.pdf"
        pdf.output(pdf_output)

        print("PDF généré avec succès.")

        # Ouvrir le PDF avec le lecteur par défaut
        import os
        os.startfile(pdf_output)  

    def Selectionner(self,item):
        self.buton_imprimer.setVisible(True)
        self.buton_effacer.setVisible(True)
        self.buton_excel.setVisible(True)

        self.buton_imprimer.setEnabled(True)
        self.buton_excel.setEnabled(True)
        self.buton_effacer.setEnabled(True)


    
        row = item.row()
        col = item.column()
        valeur = item.text()
        self.id = self.tableau.item(row,0).text() 

        print(f"{row,col,valeur,id}")        
        modif = "SELECT * FROM fiches WHERE rowid = ?"
        data = db.execute(modif, (self.id,))
        resultats = data.fetchall()
        print(resultats)

    def print(self):
        donnees_tableau = []
        for row in range(self.tableau.rowCount()):
            ligne = []
            for col in range(self.tableau.columnCount()):
                item = self.tableau.item(row, col)
                if item is not None:
                    ligne.append(item.text())
                else:
                    ligne.append("")
            donnees_tableau.append(ligne)

        df = pd.DataFrame(donnees_tableau, columns=["ID", "Numéro", "Date", "Nom", "Poste", "Congé", "Évaluation"])

        excel_output = "Fenetres.xlsx"
        df.to_excel(excel_output, index=False)

        print("Fichier Excel généré avec succès :", excel_output)
        Excel = "C:\Program Files\Microsoft Office\Office16\excel.exe"
        subprocess.Popen([Excel, "Fenetres.xlsx"])

    
        
    def Effacer(self):
        ligne = self.tableau.selectedItems()
        conn = sqlite3.connect("gestion.db")
        cu = conn.cursor()

        cu.execute("DELETE FROM fiches WHERE rowid =:id",{"id":self.id})


        conn.commit()
        conn.close()
        listes = ajout_fiche()
        self.tableau.setRowCount(len(listes))
        for row,rowdata in enumerate(listes):
            for col,value in enumerate(rowdata):
                item = QTableWidgetItem(str(value))
                self.tableau.setItem(row,col,item)



    def retour_fenetre1(self):
        self.deconnecter.show()    
        self.hide()






if __name__ == "__main__":
    app = QApplication(sys.argv)
    tableau_fiche = QDialog()
    ui = Tableau_fiche_ui()
    tableau_fiche.show()
    sys.exit(app.exec_()) 